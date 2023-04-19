from copy import copy
from types import CellType
import openpyxl
from openpyxl.styles import PatternFill
import unidecode as ud


def remove_empty(ws): 
    '''Metodo que sirve para borrar todas las filas que estan vacias (no se utiliza de momento porque no sirve para borrar
       lad filas que hay que eliminar pero tienen contenido)'''
    filas = ws.max_row
    for i in range(filas, 0, -1):
        celdas_vacias = all([cell.value is None for cell in ws[i]])
        if celdas_vacias:
            ws.delete_rows(i, 1)

def unmerge_cells(ws):
    '''Metodo que desmergea las celdas de las filas a eliminar'''
    # Buscamos las celdas que estan mergeadas
    for merge in list(ws.merged_cells):
        # Separamos esas celdas mergeadas
        ws.unmerge_cells(range_string=str(merge))

def remove(ws,ini,end): 
    '''Metodo que sirve para borrar todas las filas de 1 a 11'''
    ws.delete_rows(ini,end)


def change_colour(ws): 
    '''Metodo que a todas la celdas les pone color blanco (quita el amarillo) '''
    max_row = ws.max_row
    #print('El numero de filas rellenas es: ', max_row)
    max_column = ws.max_column
    #print('El numero de columnas rellenas es: ', max_column)

    # Hacemos doble bucle para recoger todos los datos de la tabla que empezaba en fila 12 (ahora empezara en la tabla 
    # destino en la fila 1).
    for i_row in range(1, max_row + 1):
       for i_column in range(1, max_column + 1):
           cell = ws.cell(row = i_row, column = i_column)
           #If ((cell.fill.fgcolor.type == 'indexed' and cell.fill.fgcolor.indexed == 43) or
           #(cell.fill.fgcolor.type == 'rgb' and cell.fill.fgcolor.rgb == 'FFFFFF99')):
           # Cambia para todas las celdas el fondo a blanco
           cell.fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type = "solid")

def copyStyle(newCell, cell): 
    '''Metodo que copia el formato de una celda a una nueva '''
    if cell.has_style: 
        newCell.style = copy(cell.style) 
        newCell.font = copy(cell.font) 
        newCell.border = copy(cell.border) 
        newCell.fill = copy(cell.fill) 
        newCell.number_format = copy(cell.number_format) 
        newCell.protection = copy(cell.protection) 
        newCell.alignment = copy(cell.alignment)
  
def insert_column(ws, colNr):
    '''Metodo que inserta una columna nueva detrás de la columna H poniendo la cabecera y calcula el valor como G-H'''
    # Insertamos columna
    ws.insert_cols(colNr)
    #añadimos el título a la columna 
    #ws.cell(row=headerRow, column=colNr).value = headerVal

    # Damos valor G2-H2 (FTE - Role Comment 2)
    for col in ws.iter_cols(min_row=2, min_col=colNr-2, max_col=colNr-1):
        for cell in col: 
            # Tratamiento primera columna de la operacion
            cell_column1 = ws.cell(row=cell.row, column=colNr-2)
            #print('El valor de la celda de la columna 1 es: ', cell_column1.value)
            # Tratamiento segunda columna de la operacion
            cell_column2 = ws.cell(row=cell.row, column=colNr-1)
            #print('El valor de la celda de la columna 2 es: ', cell_column2.value)
            cell_resultado = ws.cell(row=cell.row, column=colNr)
            if cell_column2.value is None: 
                cell_column2.value = 0
            # Se intenta hacer la resta, si no ss posible, se indica 'Error en resta' en la celda y en rojo
            try:
                cell_resultado.value = int(cell_column1.value) - int(cell_column2.value)
            except:
                cell_resultado.value = '¡¡Error en resta¡¡'
                cell_resultado.fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type = "solid")

def format_column(ws, colNr):
    '''Metodo que da formato a la columna que se ha creado usando método range'''
    max_row = ws.max_row
    #recorremos todas las celdas de la columna G para copiar su formato a la nueva columna mediante el método
    #copyStyle definido más arriba
    for i_row in range(1, max_row + 1):
        cell_new = ws.cell(row=i_row, column= colNr)
        cell_origin = ws.cell(row=i_row, column=colNr-2)
        if cell_new.value != '¡¡Error en resta¡¡':
            copyStyle(cell_new, cell_origin)

def format_column_iter(ws, colNr):  
    '''Metodo que da formato a la columna que se ha creado usando método iter_cols'''
    #iter_cols devuelve un generador de tuplas, donde cada tupla contiene todas las celdas de una columna en particular, 
    #desde la fila inicial hasta la fila final (establecidas con los argumentos min_row y max_row, respectivamente).
    #itera a través de todas las columnas de la hoja de trabajo (ws) y 
    #selecciona solo la columna anterior a la que se desea copiar (colNr-1).
    for col in ws.iter_cols(min_row=1, min_col=colNr-1, max_col=colNr-1): 
        #recorre todas las celdas de esa columna y copia su estilo a la celda 
        #correspondiente en la nueva columna (colNr).
        #recorre todas las celdas de la columna seleccionada en la variable col, y en cada iteración, 
        #la variable cell_origin se asigna a una celda de la columna actual. 
        for cell_origin in col:
            #print("la fila a tratar es: ", cell_origin.row)
            #print("la coordenada a tratar es: ", cell_origin.coordinate)
            cell_new = ws.cell(row=cell_origin.row, column=colNr)
            copyStyle(cell_new, cell_origin)
        

def format_condition(ws, colNr, condition, color):     
    '''Metodo que da formato a la columna que se ha creado según criterio'''   
    #Si la celda de la columna F (Additional Notes) es igual a CRITICA se cambia el fondo de la celda de la
     #columna nueva (I - FTES. Pdtes.) a amarillo
    max_row = ws.max_row    
    for i_row in range(1, max_row + 1):   
        cell_new = ws.cell(row=i_row, column= colNr)
        cell_condition = ws.cell(row=i_row, column= colNr-3)
        if cell_condition.value.upper() == condition:
           cell_new.fill = color

def format_condition_iter(ws, colNr, condition, color):     
    '''Metodo que da formato a la columna que se ha creado según criterio utilizando iter_cols y enumerate'''  
    #colNr: es el número de la columna que se desea formatear.
    #condition: es el valor que se debe buscar en la columna colNr-3 para aplicar el formato condicional.
    #Si la celda de la columna F (Additional Notes) es igual a CRITICA se cambia el fondo de la celda de la
    #columna nueva (I - FTES. Pdtes.) a amarillo
    #values_only=True ->indica si solo se deben recorrer los valores de las celdas sin objetos "Cell"
    #si utilizamos values_only=True necesitamos el enumerate, si no lo utilizamos luego tendremos que acceder
    #al atributo .value
    for col in ws.iter_cols(min_row=1, min_col=colNr-3, max_col=colNr-3, values_only=True): 
        #print(col)
        #col tiene ('Additional Notes', 'CRITICA', 'URGENTE', 'URGENTE', None, 'CRITICA', ',')
        #el siguiente for itera a través de cada celda en la columna seleccionada
        #La función enumerate() devuelve una tupla con un índice i y un valor value correspondiente al 
        #valor de la celda en la posición i.
        for i, value in enumerate(col):
        #Si el valor de la celda coincide con la condición (value == condition), entonces el código cambia el colorç
        #de fondo de la celda correspondiente en la columna colNr usando el método fill de la celda.   
           if value.upper() == condition:
              #print("el valor de lo encontrado es:", value)
              #print("el valor de i es:", i)
              #en cell_coor almacenamos la coordenada de la celda en la columna correspondiente (colNr) y 
              #la fila correspondiente (i + 1). Esta coordenada se almacena en la variable cell_coor.
              #utilizamos i+1 porque aquí empezamos desde 0 y en excel se empieza desde 1
              cell_coor = ws.cell(row=i+1,column= colNr).coordinate
              #print('coordenada: ', cell_coor)
              #Se utiliza la coordenada de la celda (cell_coor) para obtener un objeto Cell de la hoja de cálculo (cell) 
              #correspondiente a la celda en la fila y columna especificadas.
              cell = ws[cell_coor]
              cell.fill = color      

def format_condition_iter2(ws, colNr, condition, color):     
    '''Metodo que da formato a la columna que se ha creado según criterio utilizando iter_cols y cell'''  
    #colNr: es el número de la columna que se desea formatear.
    #condition: es el valor que se debe buscar en la columna colNr-3 para aplicar el formato condicional.
    #Si la celda de la columna F (Additional Notes) es igual a CRITICA se cambia el fondo de la celda de la
    #columna nueva (I - FTES. Pdtes.) a amarillo
    #si no recuperamos el valor con values_only=True tendremos que acceder luego al 
    #al atributo .value
    for col_condition in ws.iter_cols(min_row=1, min_col=colNr-3, max_col=colNr-3): 
        #print(col_condition)
        #col tiene ('Additional Notes', 'CRITICA', 'URGENTE', 'URGENTE', None, 'CRITICA', ',')
        #el siguiente for itera a través de cada celda en la columna seleccionada
        for cell_condition in col_condition:
        #Si el valor de la celda tiene valor y coincide con la condición (value == condition) pasada a mayusculas,
        # entonces el código cambia el color de fondo de la celda correspondiente en la columna colNr usando 
        # el método fill de la celda.   
           if cell_condition.value and cell_condition.value.upper() == condition:
              #print('entra en if del condition con coordinada de celda:', cell_condition.coordinate)
              #print('La fila de la celda de la condicion es: ',cell_condition.row)
              #print("el valor de lo encontrado es:", cell_condition.value)
              #en cell_coor almacenamos la coordenada de la celda 
              # construyo la coordenada de la celda a la que quiero cambiar el formato, a partir de la fila
              # (misma fila que la celda de la condicion) y la columna a cambiar formato (pasada por parametro)
              cell_coor = ws.cell(row=cell_condition.row, column=colNr).coordinate
              #print('coordenada: ', cell_coor)
              #Se utiliza la coordenada de la celda (cell_coor) para obtener un objeto cell de la hoja de cálculo (ws) 
              #correspondiente a la celda en la fila y columna especificadas.
              cell = ws[cell_coor]
              #se cambia el color al nuevo objeto creado
              cell.fill = color     

def delete_row(ws, colNr,condition):
    '''Metodo que borra las filas de columna C 8'TeamRequestStatus') con texto 'Draft' '''  
    for col_condition in ws.iter_cols(min_row=1, min_col=colNr, max_col=colNr):
        for cell_condition in col_condition: 
            #print("el valor de la celda es:", cell_condition.value)
            if cell_condition.value and cell_condition.value.upper() == condition:
                #print('La fila de la celda de la condicion es: ',cell_condition.row)
                #print('La columna de la celda de la condicion es: ',cell_condition.column)
                # Borra la fila de la celda que ha cumplido la condicion. con 1 le indico que borre esa fila solo
                ws.delete_rows(cell_condition.row,1)    
    
def change_header(ws, colNr, headerRow, headerVal):
    '''Metodo que cambia la cabecera de una columna'''
    #Cambiamos el título a la columna 
    ws.cell(row=headerRow, column=colNr).value = headerVal

def change_values(ws, colNr, condition1):    
    '''Metodo que cambia valores y formatos de celdas segun condiciones'''  
    #Para todo el bucle de iter_cols empezamos en la fila 2 para no tener en cuenta las cabeceras
    for col_condition in ws.iter_cols(min_row=2, min_col=colNr, max_col=colNr):
        for cell_condition in col_condition: 
            #print("el valor de la celda es:", cell_condition.value)
            # PRIMER FILTRO
            # Si la celda de la columna I(posicion 9) con titulo 'FTES. Pdtes es 0 
            #   -> en col F (posicion 6) con titulo ''CRITICIDAD' poner valor 'CUBIERTA ' 
            #   -> en columna C (posicion 3) pintar en verde el fondo de columna
            if cell_condition.value == condition1:
                #print('Filtro 1 La fila de la celda de la condicion es: ',cell_condition.row)
                #print('Filtro 1 La columna de la celda de la condicion es: ',cell_condition.column)
                cell_change = ws.cell(row=cell_condition.row, column=colNr-3)
                #cell_change.value = valor
                cell_change.value = 'CUBIERTA'
                # Me posiciono en la columna C (posicion 3) para cambiar el fondo
                cell_change = ws.cell(row=cell_condition.row, column=colNr-6)
                # En la columna C (posicion 3) cambio el fondo a verde
                #cell_change.fill = color 
                cell_change.fill = PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type = "solid")
            
            if cell_condition.value!= condition1:
                #print('Filtro 2 La fila de la celda de la condicion es: ',cell_condition.row)
                #print('Filtro 2 La columna de la celda de la condicion es: ',cell_condition.column)
                # Me posiciono en la columna E (posicion 5) para aplicar luego la segunda condicion
                cell_condition2 = ws.cell(row=cell_condition.row, column=colNr-4)
                #if cell_condition2.value == condition2:
                # SEGUNDO FILTRO
                #Si columna I(posicion 9) con titulo 'FTES. Pdtes. '!= 0 
                # y la celda de la columna E(posicion 5) con titulo 'IsPositionCritical' tiene valor y es YES (pasado
                # a mayusculas) 
                #  -> en columna F (posicion 6) con titulo 'CRITICIDAD' poner valor 'CRITICA ' (sin acento) 
                #  -> en columna C (posicion 3) pintar en rojo el fondo de columna
                if cell_condition2.value and cell_condition2.value.upper() == 'YES':
                    #print('Filtro 3 La fila de la celda de la condicion es: ',cell_condition2.row)
                    #print('Filtro 3 La columna de la celda de la condicion es: ',cell_condition2.column)
                    # Me posiciono en la columna F (posicion 6) para cambiar el valor
                    cell_change = ws.cell(row=cell_condition2.row, column=colNr-3)
                    #cell_change.value = valor
                    cell_change.value = 'CRITICA'
                    # Me posiciono en la columna C (posicion 3) para cambiar el fondo
                    cell_change = ws.cell(row=cell_condition2.row, column=colNr-6)
                    # En la columna C (posicion 3) cambio el fondo a rojo
                    #cell_change.fill = color 
                    cell_change.fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type = "solid")

                else:
                    #print('Filtro 3 La fila de la celda de la condicion es: ',cell_condition2.row)
                    #print('Filtro 3 La columna de la celda de la condicion es: ',cell_condition2.column)
                    # Me posiciono en la columna F (posicion 6) para aplicar luego la tercera condicion
                    cell_change = ws.cell(row=cell_condition2.row, column=colNr-3)
                    # TERCER FILTRO
                    #Si columna I(posicion 9) con titulo 'FTES. Pdtes. '!= 0 
                    # y columna E(posicion 5) con titulo 'IsPositionCritical'=No
                    # y celda de columna F(posicion 6) con titulo 'CRITICIDAD' tiene valor y contiene 
                    # 'Critica o urgente' (pasado a mayusculas)
                    #  -> en columna F (posicion 6) con titulo 'CRITICIDAD' poner valor 'URGENTE' 
                    #  -> en columna C (posicion 3) pintar en amarillo el fondo de columna
                    cadena = cell_change.value
                    #añadimos unidecode (con un alias ud) para poder comparar cadenas con caracteres ASCII (quitar tildes)
                    if cadena and ("CRITICA" in ud.unidecode(cadena).upper() or "CRITICO" in ud.unidecode(cadena).upper()
                                              or ("URGENTE" in ud.unidecode(cadena).upper() and 
                                                                "NO" not in ud.unidecode(cadena).upper())):
                       #print("es critica ", cell_change)
                       # En la columna F (posicion 6) cambio el valor a 'URGENTE'
                       cell_change.value = 'URGENTE'
                       # Me posiciono en la columna C (posicion 3) para cambiar el fondo
                       cell_change = ws.cell(row=cell_condition2.row, column=colNr-6)
                       #cell_change.fill = color 
                       # En la columna C (posicion 3) cambio el fondo a amarillo
                       cell_change.fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type = "solid")
                    
                    # CUARTO FILTRO
                    #Si columna I(posicion 9) con titulo 'FTES. Pdtes. '!= 0 
                    # y columna E(posicion 5) con titulo 'IsPositionCritical'=No
                    # y la celda de la columna F(posicion 6) con titulo 'CRITICIDAD'! tiene valor y es 'URGENTE' (pasado
                    # a mayusculas)
                    #  -> en columna F (posicion 6) con titulo 'CRITICIDAD' poner valor 'NORMAL' 
                    #  -> dejar fondo en blanco????
                    # Me posiciono en la columna F (posicion 6) para aplicar luego la tercera condicion, empezando
                    # en la fila row + 1 para que no machaque la cabecera 'CRITICIDAD' que entraria por != 'URGENTE'
                    cell_change = ws.cell(row=cell_condition2.row, column=colNr-3)
                    if (cell_change.value and cell_change.value.upper() != 'URGENTE') or cell_change.value is None:
                        # En la columna F (posicion 6) cambio el valor a 'normal'
                        cell_change.value = 'NORMAL'
                        

         