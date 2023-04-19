#!/bin/env python
# Cedazo - Morph, Meld and Merge data - Copyright © 2023 Iwan van der Kleijn - See LICENSE.txt for conditions
#from openpyxl import Workbook
import openpyxl

from data import change_colour, change_header, change_values, delete_row, format_column, format_column_iter, format_condition, format_condition_iter, format_condition_iter2, insert_column, remove, unmerge_cells
from miargparse import parser
from openpyxl.styles import PatternFill 
#Damos la localización del fichero de entrada
#ruta_input = "C:\\Users\\lgordoga\\L99PYTHON\\SPRINT_0\\in_corto.xlsx"
#Damos la localización del fichero de salida
#ruta_output = "C:\\Users\\lgordoga\\L99PYTHON\\SPRINT_0\\out.xlsx"
# Creamos objeto wb (libro) de tipo workbook y lo cargamos con lo del excel
args = parser.parse_args()
#ruta_input = args.ruta_input
print(args)

wb = openpyxl.load_workbook(args.ruta_input)
# Creamos objeto ws (hoja), siendo la hoja activa
#ws = wb.active 
ws = wb['Retain Report']
#print('la celda A1 es:' ,ws['A1'].value)

#Metodo que desmergea las celdas de las filas a eliminar
unmerge_cells(ws)

# Metodo que sirve para borrar todas las filas de 1 a 11
#for row in ws: 
remove(ws,1,11)

# Metodo que elimina el color amarillo de todas la celdas que sean amarillas 
change_colour(ws)

#Metodo que inserta una columna nueva detrás de la columna H poniendo la cabecera y copiando el formato de la columna H
insert_column(ws, colNr=9)
#Metodo que cambia la cabecera de una columna -> Colum 'FTES. Pdtes.' a la nueva columna
change_header(ws, colNr=9, headerRow=1, headerVal='FTES. Pdtes.')

#Metodo que da formato a la columna que se ha creado usando range
format_column(ws, colNr= 9)
#Metodo que da formato a la columna que se ha creado utilizando iter_cols
#format_column_iter(ws, colNr= 9)

#Metodo que da formato a la columna que se ha creado según criterio  
color_yellow = PatternFill(start_color="FFFFFF99", end_color="FFFFFF99", fill_type = "solid")
#format_condition(ws, colNr= 9, condition="CRITICA", color=color_yellow)
#Metodo que da formato a la columna que se ha creado según criterio utilizando iter_cols y enumerate
#format_condition_iter(ws, colNr= 9, condition="CRITICA", color=color_yellow)
#Metodo que da formato a la columna que se ha creado según criterio utilizando iter_cols y cell
format_condition_iter2(ws, colNr= 9, condition="CRITICA", color=color_yellow)

# Metodo que borra las filas de columna C 8'TeamRequestStatus') con texto 'Draft' 
#for row in ws: 
delete_row(ws,colNr= 3,condition="DRAFT")

#Metodo que cambia la cabecera de una columna -> Cambiar nombre de Col F de 'Additional Notes' a 'CRITICIDAD'
change_header(ws, colNr=6, headerRow=1, headerVal='CRITICIDAD')
#Metodo que cambia la cabecera de una columna -> Cambiar nombre de Col O de 'Team Request Comment 1' a 'CLIENTE'
change_header(ws, colNr=15, headerRow=1, headerVal='CLIENTE')

#Metodo que cambia valores y formatos de celdas segun condicioes
# PRIMER FILTRO
# Si columna I(posicion 9) con titulo 'FTES. Pdtes 0 
#   -> en col F (posicion 6) con titulo ''CRITICIDAD' poner valor 'CUBIERTA ' 
#   -> en columna C (posicion 3) pintar en verde el fondo de columna
# SEGUNDO FILTRO
#-Si columna I(posicion 9) con titulo 'FTES. Pdtes. '!= 0 
# y columna E(posicion 5) con titulo 'IsPositionCritical'=YES 
#  -> en columna F (posicion 6) con titulo 'CRITICIDAD' poner valor 'CRITICA ' (sin acento) 
#  -> en columna C (posicion 3) pintar en rojo el fondo de columna
# TERCER FILTRO
#-Si columna I(posicion 9) con titulo 'FTES. Pdtes. '!= 0 
# y columna E(posicion 5) con titulo 'IsPositionCritical'=No
# y columna F(posicion 6) con titulo 'CRITICIDAD' que contenga'Critica o urgente'
#  -> en columna F (posicion 6) con titulo 'CRITICIDAD' poner valor 'URGENTE' 
#  -> en columna C (posicion 3) pintar en amarillo el fondo de columna
# CUARTO FILTRO
#-Si columna I(posicion 9) con titulo 'FTES. Pdtes. '!= 0 
# y columna E(posicion 5) con titulo 'IsPositionCritical'=No
# y columna F(posicion 6) con titulo 'CRITICIDAD'!= 'URGENTE'
#  -> en columna F (posicion 6) con titulo 'CRITICIDAD' poner valor 'MORMAL' 
#  -> dejar fondo en blanco????
change_values(ws, colNr=9,condition1=0)

# Save the workbook to the output file
wb.save(args.ruta_output)


